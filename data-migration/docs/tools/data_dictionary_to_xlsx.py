"""Build the DEMOS data-dictionary Excel workbook.

Run from the repo root:

    uv run --with openpyxl python docs/tools/data_dictionary_to_xlsx.py \\
      reports/inputs/demos_data_dictionary.xlsx

Schema sources
--------------
The DEMOS app team owns the entire ``demos_app`` shape -- parents,
FKs, indexes, sequences, *and* the ``*_history`` tables plus the
``revision_type_enum`` -- produced by Prisma and applied by
``migrate ddl`` from the artifact pinned in
``reports/prisma_ddl.sha256``. The migration repo only owns the JSONB
schema registry (``sql/01_ddl_supplements/00_jsonb_schema_registry.sql``
-- ``migration.jsonb_schemas`` plus the JSONB validation trigger
function and the ``revalidate_jsonb`` helper).

The workbook needs the parent shape, the migration-private registry,
and the Prisma-owned history tables, so the generator reads three
sources:

1. The hand-authored DEMOS data model: a single mermaid file at
   ``../demos/data/docs/DEMOS_Data_Model.mmd``. Each entity is a
   ``name:::class { ... }`` block whose ``:::class`` marks it as a
   static constraint / type limiter / data table / associative table.
   It is the canonical description of every parent table. Override the
   path with ``--mmd-file <path>`` or the ``DEMOS_DATA_MODEL_MMD`` env
   var if it lives elsewhere.
2. ``sql/01_ddl_supplements/00_jsonb_schema_registry.sql`` for the
   ``migration.jsonb_schemas`` housekeeping table.
3. The pinned, cached Prisma artifact under
   ``state/prisma_ddl/<pin>.sql`` for the DEMOS ``*_history`` tables
   (DEMOS audit shape: ``revision_id``/``revision_type``/``modified_at``
   then the mirrored parent columns). Override with
   ``--prisma-artifact <path>``.

Output structure
----------------
* ``Overview`` -- one row per table with class, history-table flag,
  primary key, column count, source, and a short purpose blurb.
* One sheet per table -- column-level detail (name, type, null,
  key/default, notes), the table's CHECK constraints with their
  English descriptions, the trigger list (kind + description), and a
  ``History (audit) table`` section reproducing the matching
  Prisma-owned ``*_history`` table when one exists.
"""

from __future__ import annotations

import argparse
import os
import re
from collections.abc import Iterable
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from schema_model import DEMOS_DATA_MODEL_MMD, MERMAID_CLASS_TO_DIR

REPO_ROOT = Path(__file__).resolve().parent.parent.parent
SUPPLEMENTS_DIR = REPO_ROOT / "sql" / "01_ddl_supplements"
DEFAULT_MMD_FILE = Path(os.environ.get("DEMOS_DATA_MODEL_MMD", str(DEMOS_DATA_MODEL_MMD)))
_WORKSPACE_ROOT = REPO_ROOT.parent


def _confined_path(raw: str, *, what: str) -> Path:
    """Resolve ``raw`` and refuse paths escaping the workspace (CWE-23 guard).

    Inputs/outputs legitimately span this repo and its sibling checkouts (e.g.
    ``mmd_sql_compare/``), so the boundary is the workspace root, not the repo.
    Anything resolving outside it (``/etc``, ``/tmp``, ``../..`` escapes) is
    rejected rather than read or written.
    """
    resolved = Path(raw).resolve()
    if not resolved.is_relative_to(_WORKSPACE_ROOT):
        raise SystemExit(f"refusing {what} outside {_WORKSPACE_ROOT}: {resolved}")
    return resolved

CLASS_DIR_TO_LABEL: dict[str, str] = {
    "static_constraints": "Static constraint",
    "type_limiters": "Type limiter",
    "data_tables": "Data table",
    "associative_tables": "Associative table",
}

# Hand-curated table-purpose blurbs live in table_purpose.py so the
# AsciiDoc and XLSX data-dictionary generators share one source.
from table_purpose import (
    PURPOSE,
    STATIC_CONSTRAINT_PURPOSE,
    TYPE_LIMITER_PURPOSE,
)

# JSONB columns whose shape is enforced by the CONSTRAINT TRIGGER attached
# in sql/31_constraint_triggers/00_jsonb_validation.sql.
JSONB_TRIGGERS: dict[tuple[str, str], str] = {
    ("budget_neutrality_workbook", "validation_data"): "budget_neutrality",
    ("uipath_result", "response"): "uipath_response",
    ("uipath_value", "token_list"): "uipath_response",
}


# ---------------------------------------------------------------------------
# Mermaid parsing
# ---------------------------------------------------------------------------

# SQL types we expect to see in the mermaid model. The mmd uses short forms
# (`bool`, `double`, `numeric(N,N)`); we render the canonical Postgres form
# in the workbook for clarity.
TYPE_MAP: dict[str, str] = {
    "bool": "boolean",
    "boolean": "boolean",
    "double": "double precision",
    "smallint": "smallint",
    "integer": "integer",
    "bigint": "bigint",
    "numeric": "numeric",
    "uuid": "uuid",
    "text": "text",
    "timestamptz": "timestamptz",
    "date": "date",
    "jsonb": "jsonb",
}

NON_COLUMN_PREFIXES = (
    "CONSTRAINT_CHECK",
    "CONSTRAINT_TRIGGER",
    "TRIGGER_BEFORE",
    "TRIGGER_AFTER",
    "PARTIAL_UNIQUE_INDEX",
    "NULLS_NOT_DISTINCT_UNIQUE_INDEX",
    "HAS_HISTORY",
)

_TABLE_BLOCK_RE = re.compile(
    r"^[ \t]*(?P<name>[A-Za-z_][A-Za-z_0-9]*):::(?P<cls>\w+)[ \t]*\{(?P<body>.*?)^[ \t]*\}",
    re.MULTILINE | re.DOTALL,
)

# A column line: leading whitespace, then a TYPE token (possibly numeric(N,N)),
# then the column name, then optional flag list (PK/FK/UK with commas), then
# an optional double-quoted description.
_COLUMN_LINE_RE = re.compile(
    r"""
    ^[ \t]*
    (?P<type>[A-Za-z_]+(?:\([^)]+\))?)
    \s+
    (?P<name>[A-Za-z_][A-Za-z_0-9]*)
    (?:\s+(?P<flags>[A-Z]+(?:\s*,\s*[A-Z]+)*))?
    (?:\s+"(?P<desc>(?:\\.|[^"\\])*)")?
    \s*$
    """,
    re.VERBOSE,
)

# Constraint / trigger / unique-index / has_history lines.
_NON_COLUMN_LINE_RE = re.compile(
    r"""
    ^[ \t]*
    (?P<kind>CONSTRAINT_CHECK|CONSTRAINT_TRIGGER[A-Z_]*|TRIGGER_BEFORE[A-Z_]*|
              TRIGGER_AFTER[A-Z_]*|PARTIAL_UNIQUE_INDEX|NULLS_NOT_DISTINCT_UNIQUE_INDEX)
    \s+
    (?P<name>[A-Za-z_][A-Za-z_0-9]*)
    (?:\s+"(?P<desc>(?:\\.|[^"\\])*)")?
    \s*$
    """,
    re.VERBOSE,
)

# A description line may carry several "<br>"-separated annotations:
#   FK###: <columns> ∈ <target> [DEFERRED]
#   UK: <columns>
#   NULLABLE
_FK_RE = re.compile(
    r"FK\d+:\s*(?P<src>[^∈]+?)\s*∈\s*(?P<tgt>[^<]+?)(?:\s+(?P<deferred>DEFERRED))?\s*$"
)
_UK_RE = re.compile(r"UK:\s*(?P<cols>.+?)\s*$")


@dataclass
class FK:
    src_cols: str
    target: str  # e.g. "demonstration.id" or "application.(id, application_type_id)"
    deferred: bool = False


@dataclass
class Column:
    name: str
    sql_type: str
    flags: set[str] = field(default_factory=set)
    nullable: bool = False
    fks: list[FK] = field(default_factory=list)
    unique_groups: list[str] = field(default_factory=list)


@dataclass
class CheckConstraint:
    name: str
    description: str


@dataclass
class IndexConstraint:
    name: str
    description: str
    kind: str = "PARTIAL UNIQUE INDEX"


@dataclass
class Trigger:
    kind: str
    name: str
    description: str


@dataclass
class TableSpec:
    name: str
    cls: str  # mermaid class: staticConstraint / typeLimiter / dataTable / associativeTable
    cls_dir: str  # source directory: static_constraints / type_limiters / ...
    schema: str = "demos_app"
    columns: list[Column] = field(default_factory=list)
    checks: list[CheckConstraint] = field(default_factory=list)
    triggers: list[Trigger] = field(default_factory=list)
    partial_indexes: list[IndexConstraint] = field(default_factory=list)
    has_history: bool = False


def _split_desc(desc: str) -> list[str]:
    if not desc:
        return []
    return [seg.strip() for seg in desc.split("<br>") if seg.strip()]


def _parse_column(match: re.Match[str]) -> Column:
    flags: set[str] = set()
    if match.group("flags"):
        flags = {tok.strip() for tok in match.group("flags").split(",") if tok.strip()}
    desc = match.group("desc") or ""

    sql_type_raw = match.group("type")
    base = sql_type_raw.split("(", 1)[0].lower()
    sql_type = TYPE_MAP.get(base, sql_type_raw)
    if "(" in sql_type_raw and base in {"numeric", "decimal"}:
        # Preserve precision/scale when present.
        sql_type = sql_type_raw.lower()

    col = Column(name=match.group("name"), sql_type=sql_type, flags=flags)

    for seg in _split_desc(desc):
        if seg.upper() == "NULLABLE":
            col.nullable = True
            continue
        m_fk = _FK_RE.match(seg)
        if m_fk:
            col.fks.append(
                FK(
                    src_cols=m_fk.group("src").strip(),
                    target=m_fk.group("tgt").strip(),
                    deferred=bool(m_fk.group("deferred")),
                )
            )
            continue
        m_uk = _UK_RE.match(seg)
        if m_uk:
            col.unique_groups.append(m_uk.group("cols").strip())
            continue
        # Anything else stays unparsed; we don't currently surface it.
    return col


def _parse_table_block(name: str, cls: str, cls_dir: str, body: str) -> TableSpec:
    spec = TableSpec(name=name, cls=cls, cls_dir=cls_dir)
    for raw_line in body.splitlines():
        line = raw_line.rstrip()
        if not line.strip():
            continue
        # Non-column constructs first (HAS_HISTORY is a bare line).
        stripped = line.lstrip()
        if stripped.startswith("HAS_HISTORY"):
            spec.has_history = "true" in stripped.lower()
            continue
        m_nc = _NON_COLUMN_LINE_RE.match(line)
        if m_nc:
            kind = m_nc.group("kind")
            entry = (kind, m_nc.group("name"), m_nc.group("desc") or "")
            if kind == "CONSTRAINT_CHECK":
                spec.checks.append(CheckConstraint(entry[1], entry[2]))
            elif kind == "PARTIAL_UNIQUE_INDEX":
                spec.partial_indexes.append(
                    IndexConstraint(entry[1], entry[2], "PARTIAL UNIQUE INDEX")
                )
            elif kind == "NULLS_NOT_DISTINCT_UNIQUE_INDEX":
                spec.partial_indexes.append(
                    IndexConstraint(entry[1], entry[2], "UNIQUE INDEX (NULLS NOT DISTINCT)")
                )
            else:
                spec.triggers.append(Trigger(*entry))
            continue
        # Anything else must be a column.
        m_col = _COLUMN_LINE_RE.match(line)
        if m_col:
            spec.columns.append(_parse_column(m_col))
            continue
        # Fallthrough: skip silently. The split_and_compare.py script
        # treats unknown lines the same way.
    return spec


def parse_mmd_file(mmd_file: Path) -> list[TableSpec]:
    """Parse every table block from the single DEMOS data model file.

    Each block is ``name:::mermaidClass { ... }``; the mermaid class is mapped
    to the class-directory label the workbook groups on. Blocks with any other
    class (e.g. the diagram ``legend``) are skipped.
    """
    out: list[TableSpec] = []
    text = mmd_file.read_text(encoding="utf-8")
    for m in _TABLE_BLOCK_RE.finditer(text):
        cls_dir = MERMAID_CLASS_TO_DIR.get(m.group("cls"))
        if cls_dir is None:
            continue
        out.append(
            _parse_table_block(
                name=m.group("name"),
                cls=m.group("cls"),
                cls_dir=cls_dir,
                body=m.group("body"),
            )
        )
    return out


# ---------------------------------------------------------------------------
# Supplements parsing
# ---------------------------------------------------------------------------


def _strip_sql_comments(sql: str) -> str:
    return re.sub(r"--[^\n]*", "", sql)


_CREATE_RE = re.compile(
    r"CREATE\s+TABLE\s+IF\s+NOT\s+EXISTS\s+(?P<schema>\w+)\.(?P<name>\w+)\s*\(",
    re.IGNORECASE,
)


def _balanced_body(sql: str, open_idx: int) -> tuple[str, int]:
    assert sql[open_idx] == "("
    depth = 0
    in_dollar = False
    i = open_idx
    while i < len(sql):
        ch = sql[i]
        if not in_dollar and sql[i : i + 2] == "$$":
            in_dollar = True
            i += 2
            continue
        if in_dollar and sql[i : i + 2] == "$$":
            in_dollar = False
            i += 2
            continue
        if in_dollar:
            i += 1
            continue
        if ch == "(":
            depth += 1
        elif ch == ")":
            depth -= 1
            if depth == 0:
                return sql[open_idx + 1 : i], i + 1
        i += 1
    raise ValueError("Unbalanced parens in CREATE TABLE body")


def _split_top_level(body: str) -> list[str]:
    parts: list[str] = []
    depth = 0
    buf: list[str] = []
    for ch in body:
        if ch == "(":
            depth += 1
            buf.append(ch)
        elif ch == ")":
            depth -= 1
            buf.append(ch)
        elif ch == "," and depth == 0:
            parts.append("".join(buf).strip())
            buf = []
        else:
            buf.append(ch)
    tail = "".join(buf).strip()
    if tail:
        parts.append(tail)
    return parts


@dataclass
class SqlColumn:
    name: str
    sql_type: str
    not_null: bool
    default: str
    inline_pk: bool
    inline_check: str


@dataclass
class SqlTable:
    schema: str
    name: str
    columns: list[SqlColumn] = field(default_factory=list)
    primary_key_cols: list[str] = field(default_factory=list)


def _parse_sql_column(item: str) -> SqlColumn:
    tokens = item.split()
    name = tokens[0]
    rest = tokens[1:]

    type_parts: list[str] = []
    i = 0
    while i < len(rest):
        upper = rest[i].upper()
        if upper in {"NOT", "NULL", "DEFAULT", "PRIMARY", "UNIQUE", "CHECK", "REFERENCES"}:
            break
        type_parts.append(rest[i])
        i += 1
    sql_type = " ".join(type_parts)

    not_null = False
    default = ""
    inline_pk = False
    inline_check = ""
    while i < len(rest):
        upper = rest[i].upper()
        if upper == "NOT" and i + 1 < len(rest) and rest[i + 1].upper() == "NULL":
            not_null = True
            i += 2
            continue
        if upper == "DEFAULT":
            j = i + 1
            depth = 0
            buf: list[str] = []
            while j < len(rest):
                u = rest[j].upper()
                if depth == 0 and u in {"NOT", "PRIMARY", "UNIQUE", "CHECK", "REFERENCES"}:
                    break
                buf.append(rest[j])
                depth += rest[j].count("(") - rest[j].count(")")
                j += 1
            default = " ".join(buf).strip().rstrip(",")
            i = j
            continue
        if upper == "PRIMARY" and i + 1 < len(rest) and rest[i + 1].upper() == "KEY":
            inline_pk = True
            i += 2
            continue
        if upper == "CHECK":
            j = i + 1
            depth = 0
            buf2: list[str] = []
            started = False
            while j < len(rest):
                tok = rest[j]
                depth += tok.count("(") - tok.count(")")
                buf2.append(tok)
                if "(" in tok:
                    started = True
                if started and depth == 0:
                    j += 1
                    break
                j += 1
            inline_check = " ".join(buf2).strip()
            i = j
            continue
        i += 1
    return SqlColumn(
        name=name,
        sql_type=sql_type,
        not_null=not_null,
        default=default,
        inline_pk=inline_pk,
        inline_check=inline_check,
    )


def _parse_sql_table_body(body: str) -> tuple[list[SqlColumn], list[str]]:
    cols: list[SqlColumn] = []
    pk_cols: list[str] = []
    for item in _split_top_level(body):
        s = item.strip()
        if not s:
            continue
        upper = s.upper()
        if upper.startswith("CONSTRAINT "):
            m = re.match(
                r"CONSTRAINT\s+\w+\s+PRIMARY\s+KEY\s*\((?P<cols>[^)]+)\)",
                s,
                re.IGNORECASE,
            )
            if m:
                pk_cols = [c.strip() for c in m.group("cols").split(",")]
            continue
        if upper.startswith("PRIMARY KEY"):
            m = re.match(r"PRIMARY\s+KEY\s*\((?P<cols>[^)]+)\)", s, re.IGNORECASE)
            if m:
                pk_cols = [c.strip() for c in m.group("cols").split(",")]
            continue
        cols.append(_parse_sql_column(s))
    if not pk_cols:
        pk_cols = [c.name for c in cols if c.inline_pk]
    return cols, pk_cols


def parse_supplement_tables(path: Path) -> list[SqlTable]:
    raw = _strip_sql_comments(path.read_text(encoding="utf-8"))
    out: list[SqlTable] = []
    pos = 0
    while True:
        m = _CREATE_RE.search(raw, pos)
        if not m:
            break
        schema = m.group("schema")
        name = m.group("name")
        body, end_idx = _balanced_body(raw, m.end() - 1)
        cols, pks = _parse_sql_table_body(body)
        out.append(SqlTable(schema=schema, name=name, columns=cols, primary_key_cols=pks))
        pos = end_idx
    return out


# ---------------------------------------------------------------------------
# Prisma artifact parsing (DEMOS-owned *_history tables)
# ---------------------------------------------------------------------------

# Prisma emits quoted, schema-unqualified DDL: `CREATE TABLE "amendment_history" (`.
_PRISMA_CREATE_RE = re.compile(
    r'CREATE\s+TABLE\s+(?:IF\s+NOT\s+EXISTS\s+)?"(?P<name>[A-Za-z_][A-Za-z_0-9]*)"\s*\(',
    re.IGNORECASE,
)
_PRISMA_PK_RE = re.compile(
    r'(?:CONSTRAINT\s+"[^"]+"\s+)?PRIMARY\s+KEY\s*\((?P<cols>[^)]+)\)',
    re.IGNORECASE,
)


def _unquote(token: str) -> str:
    return token.strip().strip('"')


def _parse_prisma_column(item: str) -> SqlColumn | None:
    s = item.strip().rstrip(",").strip()
    if not s:
        return None
    m = re.match(r'"(?P<name>[^"]+)"\s+(?P<rest>.+)$', s, re.DOTALL)
    if not m:
        return None
    name = m.group("name")
    rest = m.group("rest").strip()
    rest_up = (" " + rest.upper())
    cut = len(rest)
    for kw in (" NOT NULL", " NULL", " DEFAULT ", " PRIMARY KEY"):
        idx = rest_up.find(kw)
        if idx != -1:
            cut = min(cut, idx - 1)  # -1 to drop the leading space we prepended
    sql_type = _unquote(rest[: max(cut, 0)].strip())
    not_null = " NOT NULL" in rest_up
    inline_pk = "PRIMARY KEY" in rest_up
    default = ""
    m_def = re.search(
        r"DEFAULT\s+(?P<val>.+?)(?:\s+NOT\s+NULL|\s+PRIMARY\s+KEY|\s*$)",
        rest,
        re.IGNORECASE,
    )
    if m_def:
        default = m_def.group("val").strip().rstrip(",")
    return SqlColumn(
        name=name,
        sql_type=sql_type,
        not_null=not_null,
        default=default,
        inline_pk=inline_pk,
        inline_check="",
    )


def parse_prisma_history_tables(path: Path) -> list[SqlTable]:
    """Extract the DEMOS-owned ``*_history`` tables from the Prisma artifact."""
    raw = _strip_sql_comments(path.read_text(encoding="utf-8"))
    out: list[SqlTable] = []
    pos = 0
    while True:
        m = _PRISMA_CREATE_RE.search(raw, pos)
        if not m:
            break
        name = m.group("name")
        body, end_idx = _balanced_body(raw, m.end() - 1)
        pos = end_idx
        if not name.endswith("_history"):
            continue
        cols: list[SqlColumn] = []
        pk_cols: list[str] = []
        for item in _split_top_level(body):
            s = item.strip()
            if not s:
                continue
            upper = s.upper()
            if "PRIMARY KEY" in upper and (
                upper.startswith("CONSTRAINT") or upper.startswith("PRIMARY KEY")
            ):
                m_pk = _PRISMA_PK_RE.search(s)
                if m_pk:
                    pk_cols = [_unquote(c) for c in m_pk.group("cols").split(",")]
                continue
            if upper.startswith("CONSTRAINT") or upper.startswith("FOREIGN KEY"):
                continue
            col = _parse_prisma_column(s)
            if col is not None:
                cols.append(col)
        for c in cols:
            if c.name in pk_cols:
                c.inline_pk = True
        out.append(
            SqlTable(
                schema="demos_app", name=name, columns=cols, primary_key_cols=pk_cols
            )
        )
    return out


# ---------------------------------------------------------------------------
# Workbook assembly
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="D9E2F3")
HEADER_FONT = Font(bold=True)
SECTION_FILL = PatternFill("solid", fgColor="FFE699")
SECTION_FONT = Font(bold=True, italic=True)
NOTE_FONT = Font(italic=True, color="555555")
WRAP = Alignment(wrap_text=True, vertical="top")


def _short_tab_name(name: str, used: set[str]) -> str:
    if len(name) <= 31 and name not in used:
        return name
    abbrevs = [
        ("budget_neutrality", "bn"),
        ("demonstration", "demo"),
        ("application", "app"),
        ("deliverable", "dlv"),
        ("extension", "ext"),
        ("assignment", "asgn"),
        ("configuration", "cfg"),
        ("permission", "perm"),
        ("suggestion", "sug"),
    ]
    candidate = name
    for full, short in abbrevs:
        if len(candidate) <= 31 and candidate not in used:
            break
        candidate = candidate.replace(full, short)
    candidate = candidate[:31]
    if candidate in used:
        i = 1
        while True:
            suffix = f"_{i}"
            trimmed = candidate[: 31 - len(suffix)] + suffix
            if trimmed not in used:
                candidate = trimmed
                break
            i += 1
    return candidate


def _set_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_header(ws, row: int, headers: list[str]) -> None:
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP


def _write_section(ws, row: int, label: str, span: int) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    cell = ws.cell(row=row, column=1, value=label)
    cell.fill = SECTION_FILL
    cell.font = SECTION_FONT


def _qualify(target: str) -> str:
    """Prefix a bare table name with `demos_app.`. Leave qualified targets alone."""
    return target if "." in target.split("(", 1)[0] else f"demos_app.{target}"


def _key_default_cell(spec: TableSpec, col: Column) -> str:
    bits: list[str] = []
    if "PK" in col.flags:
        bits.append("PK")
    if "UK" in col.flags:
        bits.append("UK")
    for fk in col.fks:
        bits.append(f"FK -> {_qualify(fk.target)}{' (DEFERRED)' if fk.deferred else ''}")
    return "; ".join(bits)


def _notes_cell(spec: TableSpec, col: Column) -> str:
    notes: list[str] = []
    schema = JSONB_TRIGGERS.get((spec.name, col.name))
    if schema:
        notes.append(f"JSONB shape -> migration.jsonb_schemas.{schema}")
    # Surface unique-key groupings only when the column is part of a
    # multi-column UK; the single-column form is already covered by the UK
    # flag in the Key column.
    multi_uks = [g for g in col.unique_groups if "," in g]
    if multi_uks:
        formatted = [g if g.startswith("(") else f"({g})" for g in multi_uks]
        notes.append("UK group: " + " | ".join(formatted))
    return "; ".join(notes)


def _purpose_for(spec: TableSpec) -> str:
    if spec.name in PURPOSE:
        return PURPOSE[spec.name]
    if spec.cls_dir == "static_constraints":
        return STATIC_CONSTRAINT_PURPOSE
    if spec.cls_dir == "type_limiters":
        return TYPE_LIMITER_PURPOSE
    return ""


def _format_pk(spec: TableSpec) -> str:
    pk_cols = [c.name for c in spec.columns if "PK" in c.flags]
    return f"({', '.join(pk_cols)})" if pk_cols else ""


def _group_label(cls_dir: str) -> str:
    return CLASS_DIR_TO_LABEL.get(cls_dir, cls_dir)


def _null_str(col: Column) -> str:
    if "PK" in col.flags:
        return "N"
    return "Y" if col.nullable else "N"


def _pin_value(pin_path: Path) -> str:
    if not pin_path.exists():
        return "<missing>"
    return pin_path.read_text(encoding="utf-8").strip().split()[0]


def write_overview(
    wb: Workbook,
    primary: list[tuple[str, TableSpec | SqlTable]],
    history_names: set[str],
    pin: str,
) -> None:
    ws = wb.create_sheet("Overview", 0)
    ws.freeze_panes = "A4"
    ws.cell(row=1, column=1, value="DEMOS data dictionary").font = Font(
        bold=True, size=14
    )
    src_note = (
        "Source: hand-authored DEMOS data model "
        "../demos/data/docs/DEMOS_Data_Model.mmd (parents) + "
        "sql/01_ddl_supplements/ (migration.jsonb_schemas) + the pinned "
        "Prisma artifact (DEMOS-owned *_history tables). "
        f"Prisma DDL pin = {pin}."
    )
    ws.cell(row=2, column=1, value=src_note).font = NOTE_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)

    headers = [
        "Tab",
        "Table",
        "Schema",
        "Class",
        "Has history table",
        "Primary key",
        "Columns",
        "Purpose",
    ]
    _write_header(ws, 3, headers)
    _set_widths(ws, [28, 38, 12, 22, 18, 38, 9, 80])

    row_idx = 4
    for tab_name, tbl in primary:
        ws.cell(row=row_idx, column=1, value=tab_name).hyperlink = (
            f"#'{tab_name}'!A1"
        )
        ws.cell(row=row_idx, column=1).style = "Hyperlink"
        if isinstance(tbl, TableSpec):
            ws.cell(row=row_idx, column=2, value=tbl.name)
            ws.cell(row=row_idx, column=3, value=tbl.schema)
            ws.cell(row=row_idx, column=4, value=_group_label(tbl.cls_dir))
            ws.cell(
                row=row_idx,
                column=5,
                value="yes" if tbl.name in history_names else "no",
            )
            ws.cell(row=row_idx, column=6, value=_format_pk(tbl))
            ws.cell(row=row_idx, column=7, value=len(tbl.columns))
            ws.cell(row=row_idx, column=8, value=_purpose_for(tbl)).alignment = WRAP
        else:  # SqlTable -- migration.jsonb_schemas
            ws.cell(row=row_idx, column=2, value=tbl.name)
            ws.cell(row=row_idx, column=3, value=tbl.schema)
            ws.cell(row=row_idx, column=4, value="Migration housekeeping")
            ws.cell(row=row_idx, column=5, value="no")
            ws.cell(
                row=row_idx,
                column=6,
                value=f"({', '.join(tbl.primary_key_cols)})" if tbl.primary_key_cols else "",
            )
            ws.cell(row=row_idx, column=7, value=len(tbl.columns))
            ws.cell(
                row=row_idx,
                column=8,
                value=PURPOSE.get(tbl.name, ""),
            ).alignment = WRAP
        row_idx += 1


def _history_section(ws, start_row: int, history: SqlTable) -> int:
    row = start_row
    _write_section(
        ws, row, f"History (audit) table: {history.schema}.{history.name}", 5
    )
    row += 1
    _write_header(ws, row, ["Column", "Type", "Null", "Key / Default", "Notes"])
    row += 1
    for c in history.columns:
        ws.cell(row=row, column=1, value=c.name)
        ws.cell(row=row, column=2, value=c.sql_type)
        nn = c.not_null or c.inline_pk
        ws.cell(row=row, column=3, value="N" if nn else "Y")
        kbits: list[str] = []
        if c.inline_pk:
            kbits.append("PK")
        if c.default:
            kbits.append(f"DEFAULT {c.default}")
        ws.cell(row=row, column=4, value="; ".join(kbits)).alignment = WRAP
        ws.cell(row=row, column=5, value=c.inline_check).alignment = WRAP
        row += 1
    return row


def _migration_table_sheet(
    wb: Workbook, tab_name: str, tbl: SqlTable
) -> None:
    ws = wb.create_sheet(tab_name)
    ws.freeze_panes = "A4"
    ws.cell(row=1, column=1, value=f"{tbl.schema}.{tbl.name}").font = Font(
        bold=True, size=14
    )
    ws.cell(row=2, column=1, value=PURPOSE.get(tbl.name, "")).alignment = WRAP
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
    _write_header(ws, 4, ["Column", "Type", "Null", "Key / Default", "Notes"])
    _set_widths(ws, [32, 22, 6, 60, 80])
    row = 5
    for c in tbl.columns:
        ws.cell(row=row, column=1, value=c.name)
        ws.cell(row=row, column=2, value=c.sql_type)
        nn = c.not_null or c.inline_pk
        ws.cell(row=row, column=3, value="N" if nn else "Y")
        kbits: list[str] = []
        if c.inline_pk:
            kbits.append("PK")
        if c.default:
            kbits.append(f"DEFAULT {c.default}")
        ws.cell(row=row, column=4, value="; ".join(kbits)).alignment = WRAP
        ws.cell(row=row, column=5, value=c.inline_check).alignment = WRAP
        row += 1


def write_table_sheet(
    wb: Workbook,
    tab_name: str,
    spec: TableSpec,
    history_by_parent: dict[str, SqlTable],
) -> None:
    ws = wb.create_sheet(tab_name)
    ws.freeze_panes = "A4"

    ws.cell(row=1, column=1, value=f"{spec.schema}.{spec.name}").font = Font(
        bold=True, size=14
    )
    purpose = _purpose_for(spec)
    if purpose:
        ws.cell(row=2, column=1, value=purpose).alignment = WRAP
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)

    headers = ["Column", "Type", "Null", "Key / Default", "Notes"]
    _write_header(ws, 4, headers)
    _set_widths(ws, [34, 22, 6, 56, 60])

    row = 5
    for col in spec.columns:
        ws.cell(row=row, column=1, value=col.name)
        ws.cell(row=row, column=2, value=col.sql_type)
        ws.cell(row=row, column=3, value=_null_str(col))
        ws.cell(row=row, column=4, value=_key_default_cell(spec, col)).alignment = WRAP
        ws.cell(row=row, column=5, value=_notes_cell(spec, col)).alignment = WRAP
        row += 1

    if spec.checks or spec.partial_indexes:
        row += 1
        _write_section(ws, row, "Checks and partial indexes", 5)
        row += 1
        _write_header(ws, row, ["Name", "Kind", "Description", "", ""])
        row += 1
        for chk in spec.checks:
            ws.cell(row=row, column=1, value=chk.name)
            ws.cell(row=row, column=2, value="CHECK")
            ws.cell(row=row, column=3, value=chk.description).alignment = WRAP
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)
            row += 1
        for idx in spec.partial_indexes:
            ws.cell(row=row, column=1, value=idx.name)
            ws.cell(row=row, column=2, value=idx.kind)
            ws.cell(row=row, column=3, value=idx.description).alignment = WRAP
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)
            row += 1

    if spec.triggers:
        row += 1
        _write_section(ws, row, "Triggers", 5)
        row += 1
        _write_header(ws, row, ["Name", "Kind", "Description", "", ""])
        row += 1
        for tg in spec.triggers:
            ws.cell(row=row, column=1, value=tg.name)
            ws.cell(row=row, column=2, value=tg.kind)
            ws.cell(row=row, column=3, value=tg.description).alignment = WRAP
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)
            row += 1

    history = history_by_parent.get(spec.name)
    if history is not None:
        row += 1
        _history_section(ws, row, history)


# ---------------------------------------------------------------------------
# Driver
# ---------------------------------------------------------------------------


def _stable_class_order(spec: TableSpec) -> tuple[int, str]:
    order = {
        "data_tables": 0,
        "associative_tables": 1,
        "static_constraints": 2,
        "type_limiters": 3,
    }
    return (order.get(spec.cls_dir, 4), spec.name)


def main(argv: Iterable[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "out_path",
        nargs="?",
        default=str(REPO_ROOT / "reports" / "inputs" / "demos_data_dictionary.xlsx"),
        help="Output xlsx path (default reports/inputs/demos_data_dictionary.xlsx).",
    )
    parser.add_argument(
        "--mmd-file",
        default=str(DEFAULT_MMD_FILE),
        help=(
            "Hand-authored DEMOS data model (single mermaid file with "
            "name:::class { ... } entity blocks). Defaults to "
            "../demos/data/docs/DEMOS_Data_Model.mmd (or $DEMOS_DATA_MODEL_MMD)."
        ),
    )
    parser.add_argument(
        "--prisma-artifact",
        default=None,
        help=(
            "Path to the cached Prisma DDL artifact carrying the DEMOS "
            "*_history tables. Defaults to state/prisma_ddl/<pin>.sql, where "
            "<pin> is read from reports/prisma_ddl.sha256."
        ),
    )
    args = parser.parse_args(argv)

    # SECURITY: confine operator-supplied paths to the workspace (CWE-23 guard).
    out_path = _confined_path(args.out_path, what="output path")
    mmd_file = _confined_path(args.mmd_file, what="mmd file")
    if not mmd_file.is_file():
        raise SystemExit(
            f"DEMOS data model not found: {mmd_file}; pass --mmd-file or set "
            "DEMOS_DATA_MODEL_MMD to the DEMOS_Data_Model.mmd path."
        )

    specs = parse_mmd_file(mmd_file)
    if not specs:
        raise SystemExit(f"no mermaid tables found in {mmd_file}")

    pin = _pin_value(REPO_ROOT / "reports" / "prisma_ddl.sha256")

    if args.prisma_artifact:
        artifact_path = _confined_path(args.prisma_artifact, what="prisma artifact")
    else:
        artifact_path = REPO_ROOT / "state" / "prisma_ddl" / f"{pin}.sql"
    if artifact_path.exists():
        history_tables = parse_prisma_history_tables(artifact_path)
    else:
        history_tables = []
        print(
            f"WARNING: Prisma artifact not found at {artifact_path}; the "
            "*_history tables (DEMOS-owned) will be omitted. Run "
            "`uv run migrate fetch-prisma` to populate the cache.",
        )
    history_by_parent: dict[str, SqlTable] = {}
    for h in history_tables:
        if h.name.endswith("_history"):
            history_by_parent[h.name[: -len("_history")]] = h

    registry_path = SUPPLEMENTS_DIR / "00_jsonb_schema_registry.sql"
    registry_tables = parse_supplement_tables(registry_path) if registry_path.exists() else []

    wb = Workbook()
    wb.remove(wb.active)

    used_tabs: set[str] = set()
    primary_tabs: list[tuple[str, TableSpec | SqlTable]] = []
    for spec in sorted(specs, key=_stable_class_order):
        tab = _short_tab_name(spec.name, used_tabs)
        used_tabs.add(tab)
        primary_tabs.append((tab, spec))

    # Migration housekeeping tables (jsonb_schemas).
    for tbl in registry_tables:
        tab = _short_tab_name(tbl.name, used_tabs)
        used_tabs.add(tab)
        primary_tabs.append((tab, tbl))

    write_overview(wb, primary_tabs, set(history_by_parent), pin)

    for tab_name, tbl in primary_tabs:
        if isinstance(tbl, TableSpec):
            write_table_sheet(wb, tab_name, tbl, history_by_parent)
        else:
            _migration_table_sheet(wb, tab_name, tbl)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(
        f"Wrote {out_path} ({len(specs)} parent tables, "
        f"{len(history_by_parent)} history tables, "
        f"{len(registry_tables)} migration tables)"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
